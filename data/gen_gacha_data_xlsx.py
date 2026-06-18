"""Generate data/ガチャデータ.xlsx — multi-sheet master data for all 3 BOXes.

Sheets (in this order):
  1. プレミアムBOX＋  (24 tsums, 2026-06 lineup)
  2. プレミアムBOX     (150 tsums, 常駐)
  3. ハピネスBOX       (14 tsums, 常駐, SL3 max)
  4. README

Skill growth columns:
  Premium/Premium+ — SL1→2, SL2→3, SL3→4, SL4→5, SL5→6 (max SL6)
  Happiness        — SL1→2, SL2→3 only (max SL3)

Sources: game8.jp と xn--bdka7fb.jp の個別ツム評価ページ
Verification: 並列調査エージェント + 2サイト・クロスチェック
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ===================================================================
# Growth patterns
# ===================================================================
PATTERN_BASES = {
    "P1": (2, 3, 4, 6),  # 2015年3月以前の常駐ツム
    "P2": (1, 2, 4, 7),  # 中堅〜21型
    "P3": (1, 2, 4, 8),  # 晩成型・近年の重め成長
    "PS": (1, 2, 2, 3),  # 特殊。ハム専用 (合計11)
}

# ハピネスBOX 専用パターン (SL3 maxなので2段階のみ)
HAPPINESS_GROWTH = (2, 4)  # SL1→2, SL2→3
HAPPINESS_PATTERN_LABEL = "HP"  # for the スプレッドシート

# ===================================================================
# 今月のツム (期間限定/月替わり) のキャラ名
# プレミアムBOX・プレミアムBOX＋の両シートで赤色塗り
# ===================================================================
LIMITED_NAMES = {"ヒーロースタイルスティッチ", "変身の達人マウイ", "ジム・ホーキンス"}

# ===================================================================
# プレミアムBOX＋ (2026-06 ラインナップ, 24体)
# ===================================================================
# (name, yomi, pattern, sl5to6, game8_url, bdka7fb_url, verified, note)
# プラスツム20体 → 全員 P3 (1,2,4,8) + sl5to6=20 → 合計36
# 新ツム4体 → P3 (1,2,4,8) + sl5to6=16または20
PLUS_TSUMS = [
    # ユーザー指定順 (Excelの備考列で番号付け済, 2026-06)
    # 1〜19 がプラスツム, 20〜23 が新ツム, 24 が ラーヤ＋
    ("ミッキー＋", "ミッキープラス", "P3", 20, "https://game8.jp/tsumtsum/676184", "https://xn--bdka7fb.jp/175393.html", "VERIFIED", ""),
    ("ドナルド＋", "ドナルドプラス", "P3", 20, "https://game8.jp/tsumtsum/676185", "https://xn--bdka7fb.jp/175397.html", "VERIFIED", ""),
    ("アリス＋", "アリスプラス", "P3", 20, "https://game8.jp/tsumtsum/682167", "https://xn--bdka7fb.jp/176331.html", "VERIFIED", ""),
    ("ストームトルーパー＋", "ストームトルーパープラス", "P3", 20, "https://game8.jp/tsumtsum/687550", "https://xn--bdka7fb.jp/177400.html", "VERIFIED", ""),
    ("ライトニング・マックィーン＋", "ライトニングマックィーンプラス", "P3", 20, "https://game8.jp/tsumtsum/696763", "https://xn--bdka7fb.jp/178482.html", "VERIFIED", ""),
    ("プルート＋", "プルートプラス", "P3", 20, "https://game8.jp/tsumtsum/705898", "https://xn--bdka7fb.jp/179461.html", "VERIFIED", ""),
    ("アリエル＋", "アリエルプラス", "P3", 20, "https://game8.jp/tsumtsum/710186", "https://xn--bdka7fb.jp/180672.html", "VERIFIED", ""),
    ("シンバ＋", "シンバプラス", "P3", 20, "https://game8.jp/tsumtsum/720585", "https://xn--bdka7fb.jp/181703.html", "VERIFIED", ""),
    ("クルエラ＋", "クルエラプラス", "P3", 20, "https://game8.jp/tsumtsum/729252", "https://xn--bdka7fb.jp/182881.html", "VERIFIED", ""),
    ("トレメイン夫人＋", "トレメインフジンプラス", "P3", 20, "https://game8.jp/tsumtsum/729253", "https://xn--bdka7fb.jp/182884.html", "VERIFIED", ""),
    ("グーフィー＋", "グーフィープラス", "P3", 20, "https://game8.jp/tsumtsum/737659", "https://xn--bdka7fb.jp/183974.html", "VERIFIED", ""),
    ("ウッディ＋", "ウッディプラス", "P3", 20, "https://game8.jp/tsumtsum/748450", "https://xn--bdka7fb.jp/184763.html", "VERIFIED", ""),
    ("バズ・ライトイヤー＋", "バズライトイヤープラス", "P3", 20, "https://game8.jp/tsumtsum/748451", "https://xn--bdka7fb.jp/184767.html", "VERIFIED", ""),
    ("ラプンツェル＋", "ラプンツェルプラス", "P3", 20, "https://game8.jp/tsumtsum/755657", "https://xn--bdka7fb.jp/198086.html", "VERIFIED", ""),
    ("デイジー＋", "デイジープラス", "P3", 20, "https://game8.jp/tsumtsum/755656", "https://xn--bdka7fb.jp/198090.html", "VERIFIED", ""),
    ("ソラ＋", "ソラプラス", "P3", 20, "https://game8.jp/tsumtsum/761270", "https://xn--bdka7fb.jp/199152.html", "VERIFIED", ""),
    ("メイベル＋", "メイベルプラス", "P3", 20, "https://game8.jp/tsumtsum/779470", "https://xn--bdka7fb.jp/201122.html", "VERIFIED", ""),
    ("マンダロリアン＋", "マンダロリアンプラス", "P3", 20, "https://game8.jp/tsumtsum/785069", "https://xn--bdka7fb.jp/202198.html", "VERIFIED", ""),
    ("グローグー＋", "グローグープラス", "P3", 20, "https://game8.jp/tsumtsum/785070", "https://xn--bdka7fb.jp/202201.html", "VERIFIED", ""),
    # 新ツム (プラスじゃない)
    ("ヒーロースタイルスティッチ", "ヒーロースタイルスティッチ", "P3", 20, "https://game8.jp/tsumtsum/786554", "https://xn--bdka7fb.jp/202429.html", "VERIFIED", "変身ツム"),
    ("ダグ", "ダグ", "P3", 16, "https://game8.jp/tsumtsum/786555", "https://xn--bdka7fb.jp/202431.html", "VERIFIED", ""),
    ("変身の達人マウイ", "ヘンシンノタツジンマウイ", "P3", 20, "https://game8.jp/tsumtsum/787752", "https://xn--bdka7fb.jp/202710.html", "VERIFIED", "変身ツム"),
    ("ジム・ホーキンス", "ジムホーキンス", "P3", 16, "https://game8.jp/tsumtsum/787753", "https://xn--bdka7fb.jp/202713.html", "VERIFIED", "2026-06新ツム"),
    # 24. ラーヤ＋ (game8 URLはユーザー指摘で 789307 に確定)
    ("ラーヤ＋", "ラーヤプラス", "P3", 20, "https://game8.jp/tsumtsum/789307", "https://xn--bdka7fb.jp/202925.html", "VERIFIED", ""),
]

# ===================================================================
# プレミアムBOX (常駐150体) — 既存スクリプトと同一
# ===================================================================
PREMIUM_TSUMS = [
    ("ミス・バニー", "ミスバニー", "P1", 18, "https://game8.jp/tsumtsum/18973", "VERIFIED", ""),
    ("オラフ", "オラフ", "P1", 18, "https://game8.jp/tsumtsum/18253", "VERIFIED", ""),
    ("ヤングオイスター", "ヤングオイスター", "P1", 20, "https://game8.jp/tsumtsum/18333", "VERIFIED", ""),
    ("アリス", "アリス", "P1", 20, "https://game8.jp/tsumtsum/18344", "VERIFIED", ""),
    ("白うさぎ", "シロウサギ", "P1", 16, "https://game8.jp/tsumtsum/18924", "VERIFIED", ""),
    ("チェシャ猫", "チェシャネコ", "P1", 14, "https://game8.jp/tsumtsum/26341", "VERIFIED", ""),
    ("ダンボ", "ダンボ", "P1", 18, "https://game8.jp/tsumtsum/18557", "VERIFIED", ""),
    ("ティンカー・ベル", "ティンカーベル", "P1", 19, "https://game8.jp/tsumtsum/18212", "VERIFIED", "xn--bdka7fb一覧表は16表記の異本あり、Game8複数ソース19採用"),
    ("スティッチ", "スティッチ", "P1", 18, "https://game8.jp/tsumtsum/18227", "VERIFIED", ""),
    ("スクランプ", "スクランプ", "P1", 12, "https://game8.jp/tsumtsum/18634", "VERIFIED", ""),
    ("マリー", "マリー", "P1", 14, "https://game8.jp/tsumtsum/18309", "VERIFIED", ""),
    ("レディ", "レディ", "P1", 12, "https://game8.jp/tsumtsum/18294", "VERIFIED", ""),
    ("ウッディ", "ウッディ", "P1", 18, "https://game8.jp/tsumtsum/18210", "VERIFIED", ""),
    ("バズ・ライトイヤー", "バズライトイヤー", "P1", 16, "https://game8.jp/tsumtsum/18613", "VERIFIED", ""),
    ("マイク", "マイク", "P1", 18, "https://game8.jp/tsumtsum/18974", "VERIFIED", ""),
    ("サリー", "サリー", "P1", 20, "https://game8.jp/tsumtsum/18243", "VERIFIED", "モンスターズインク"),
    ("ランドール", "ランドール", "P2", 21, "https://game8.jp/tsumtsum/28206", "VERIFIED", ""),
    ("エンジェル", "エンジェル", "P2", 21, "https://game8.jp/tsumtsum/27064", "VERIFIED", ""),
    ("サプライズエルサ", "サプライズエルサ", "P2", 21, "https://game8.jp/tsumtsum/25696", "VERIFIED", ""),
    ("バースデーアナ", "バースデーアナ", "P2", 21, "https://game8.jp/tsumtsum/25695", "VERIFIED", ""),
    ("ベル", "ベル", "P2", 20, "https://game8.jp/tsumtsum/21989", "VERIFIED", ""),
    ("野獣", "ヤジュウ", "P2", 21, "https://game8.jp/tsumtsum/22013", "VERIFIED", ""),
    ("ジェシー", "ジェシー", "P1", 18, "https://game8.jp/tsumtsum/18539", "VERIFIED", ""),
    ("ロッツォ", "ロッツォ", "P1", 20, "https://game8.jp/tsumtsum/18548", "VERIFIED", ""),
    ("ベイマックス", "ベイマックス", "P1", 18, "https://game8.jp/tsumtsum/17414", "VERIFIED", ""),
    ("スヴェン", "スヴェン", "P1", 16, "https://game8.jp/tsumtsum/26344", "VERIFIED", ""),
    ("ラプンツェル", "ラプンツェル", "P1", 18, "https://game8.jp/tsumtsum/17910", "VERIFIED", ""),
    ("パスカル", "パスカル", "P1", 16, "https://game8.jp/tsumtsum/26342", "VERIFIED", ""),
    ("アリエル", "アリエル", "P1", 18, "https://game8.jp/tsumtsum/17963", "VERIFIED", ""),
    ("フランダー", "フランダー", "P1", 16, "https://game8.jp/tsumtsum/19279", "VERIFIED", ""),
    ("セバスチャン", "セバスチャン", "P1", 16, "https://game8.jp/tsumtsum/18246", "VERIFIED", ""),
    ("エルサ", "エルサ", "P1", 18, "https://game8.jp/tsumtsum/17883", "VERIFIED", ""),
    ("アナ", "アナ", "P1", 18, "https://game8.jp/tsumtsum/18332", "VERIFIED", ""),
    ("マレフィセント", "マレフィセント", "P1", 20, "https://game8.jp/tsumtsum/26671", "VERIFIED", ""),
    ("バンビ", "バンビ", "P1", 16, "https://game8.jp/tsumtsum/18614", "VERIFIED", ""),
    ("とんすけ", "トンスケ", "P1", 18, "https://game8.jp/tsumtsum/18463", "VERIFIED", ""),
    ("フェアリー・ゴッドマザー", "フェアリーゴッドマザー", "P3", 20, "https://game8.jp/tsumtsum/63342", "VERIFIED", ""),
    ("ジュディ", "ジュディ", "P3", 16, "https://game8.jp/tsumtsum/60755", "VERIFIED", ""),
    ("ニック", "ニック", "P3", 16, "https://game8.jp/tsumtsum/60753", "VERIFIED", ""),
    ("フィニック", "フィニック", "P2", 14, "https://game8.jp/tsumtsum/60754", "VERIFIED", ""),
    ("アースラ", "アースラ", "P3", 20, "https://xn--bdka7fb.jp/14460.html", "VERIFIED", ""),
    ("女王", "ジョオウ", "P3", 20, "https://xn--bdka7fb.jp/122379.html", "VERIFIED", "白雪姫の女王"),
    ("マレフィセントドラゴン", "マレフィセントドラゴン", "P3", 20, "https://xn--bdka7fb.jp/14462.html", "VERIFIED", ""),
    ("スカー", "スカー", "P3", 20, "https://xn--bdka7fb.jp/13401.html", "VERIFIED", ""),
    ("シンバ", "シンバ", "P3", 16, "https://game8.jp/tsumtsum/47413", "VERIFIED", ""),
    ("ナラ", "ナラ", "P3", 16, "https://game8.jp/tsumtsum/47415", "VERIFIED", ""),
    ("白雪姫", "シラユキヒメ", "P3", 16, "https://xn--bdka7fb.jp/12459.html", "VERIFIED", ""),
    ("オーロラ姫", "オーロラヒメ", "P3", 16, "https://xn--bdka7fb.jp/12461.html", "VERIFIED", ""),
    ("ピノキオ", "ピノキオ", "P2", 21, "https://game8.jp/tsumtsum/36875", "VERIFIED", ""),
    ("ジーニー", "ジーニー", "P2", 21, "https://game8.jp/tsumtsum/32294", "VERIFIED", ""),
    ("アラジン", "アラジン", "P2", 21, "https://game8.jp/tsumtsum/31861", "VERIFIED", ""),
    ("ジャスミン", "ジャスミン", "P2", 21, "https://game8.jp/tsumtsum/31862", "VERIFIED", ""),
    ("トリトン王", "トリトンオウ", "P2", 21, "https://game8.jp/tsumtsum/29901", "VERIFIED", ""),
    ("マックィーン", "マックィーン", "P2", 21, "https://game8.jp/tsumtsum/28207", "VERIFIED", ""),
    ("メーター", "メーター", "P2", 21, "https://game8.jp/tsumtsum/28208", "VERIFIED", ""),
    ("レックス", "レックス", "P2", 21, "https://game8.jp/tsumtsum/28205", "VERIFIED", ""),
    ("ムーラン", "ムーラン", "P3", 16, "https://game8.jp/tsumtsum/151680", "VERIFIED", ""),
    ("ポカホンタス", "ポカホンタス", "P3", 16, "https://game8.jp/tsumtsum/151681", "VERIFIED", ""),
    ("フリン・ライダー", "フリンライダー", "P3", 16, "https://game8.jp/tsumtsum/150245", "VERIFIED", ""),
    ("マキシマス", "マキシマス", "P2", 14, "https://game8.jp/tsumtsum/150246", "VERIFIED", ""),
    ("ガストン", "ガストン", "P3", 20, "https://xn--bdka7fb.jp/34054.html", "VERIFIED", ""),
    ("ルミエール", "ルミエール", "P2", 14, "https://xn--bdka7fb.jp/34056.html", "VERIFIED", ""),
    ("ハデス", "ハデス", "P3", 20, "https://game8.jp/tsumtsum/137582", "VERIFIED", ""),
    ("ヘラクレス", "ヘラクレス", "P2", 14, "https://game8.jp/tsumtsum/137579", "VERIFIED", ""),
    ("フィリップ王子", "フィリップオウジ", "P2", 14, "https://xn--bdka7fb.jp/32701.html", "VERIFIED", ""),
    ("モアナ", "モアナ", "P2", 14, "https://xn--bdka7fb.jp/32297.html", "VERIFIED", ""),
    ("マウイ", "マウイ", "P2", 14, "https://game8.jp/tsumtsum/133253", "VERIFIED", ""),
    ("フック船長", "フックセンチョウ", "P3", 20, "https://xn--bdka7fb.jp/21902.html", "VERIFIED", ""),
    ("クルエラ", "クルエラ", "P3", 20, "https://xn--bdka7fb.jp/21899.html", "VERIFIED", ""),
    ("ジャファー", "ジャファー", "P3", 20, "https://xn--bdka7fb.jp/21883.html", "VERIFIED", ""),
    ("クラッシュ", "クラッシュ", "P3", 16, "https://game8.jp/tsumtsum/79981", "VERIFIED", ""),
    ("ニモ", "ニモ", "P3", 16, "https://game8.jp/tsumtsum/79396", "VERIFIED", ""),
    ("ドリー", "ドリー", "P3", 16, "https://game8.jp/tsumtsum/79395", "VERIFIED", ""),
    ("ハートの女王", "ハートノジョオウ", "P3", 20, "https://game8.jp/tsumtsum/74380", "VERIFIED", ""),
    ("マッドハッター", "マッドハッター", "P3", 16, "https://game8.jp/tsumtsum/72384", "VERIFIED", ""),
    ("シンデレラ", "シンデレラ", "P3", 16, "https://xn--bdka7fb.jp/16567.html", "VERIFIED", ""),
    ("ガントゥ", "ガントゥ", "P3", 20, "https://xn--bdka7fb.jp/108693.html", "VERIFIED", ""),
    ("フリック", "フリック", "P3", 16, "https://game8.jp/tsumtsum/324404", "VERIFIED", ""),
    ("レミー", "レミー", "P3", 16, "https://game8.jp/tsumtsum/323464", "VERIFIED", ""),
    ("ラジャー", "ラジャー", "P2", 14, "https://game8.jp/tsumtsum/320812", "VERIFIED", ""),
    ("クリストフ", "クリストフ", "P2", 14, "https://game8.jp/tsumtsum/307032", "VERIFIED", ""),
    ("ティモン", "ティモン", "P2", 14, "https://xn--bdka7fb.jp/93126.html", "VERIFIED", ""),
    ("ラフィキ", "ラフィキ", "P3", 16, "https://xn--bdka7fb.jp/93124.html", "VERIFIED", ""),
    ("ハム", "ハム", "PS", 3, "https://xn--bdka7fb.jp/90948.html", "VERIFIED", "特殊パターン(合計11、最軽量)"),
    ("ゴーテル", "ゴーテル", "P2", 14, "https://game8.jp/tsumtsum/263751", "VERIFIED", ""),
    ("ダッチェス", "ダッチェス", "P3", 16, "https://game8.jp/tsumtsum/260306", "VERIFIED", ""),
    ("ロビン・フッド", "ロビンフッド", "P2", 14, "https://xn--bdka7fb.jp/85120.html", "VERIFIED", ""),
    ("フラワー", "フラワー", "P2", 14, "https://game8.jp/tsumtsum/259942", "VERIFIED", ""),
    ("ブー", "ブー", "P2", 14, "https://game8.jp/tsumtsum/231285", "VERIFIED", ""),
    ("王子", "オウジ", "P2", 14, "https://game8.jp/tsumtsum/187409", "VERIFIED", "白雪姫の王子"),
    ("トランプ", "トランプ", "P2", 14, "https://xn--bdka7fb.jp/44874.html", "VERIFIED", "わんわん物語"),
    ("ブルー・フェアリー", "ブルーフェアリー", "P3", 20, "https://xn--bdka7fb.jp/148430.html", "VERIFIED", ""),
    ("ピーター・パン", "ピーターパン", "P3", 16, "https://xn--bdka7fb.jp/44870.html", "VERIFIED", ""),
    ("パッチ", "パッチ", "P2", 14, "https://xn--bdka7fb.jp/44880.html", "VERIFIED", ""),
    ("ティモシー", "ティモシー", "P2", 14, "https://xn--bdka7fb.jp/44876.html", "VERIFIED", ""),
    ("クルーズ・ラミレス", "クルーズラミレス", "P3", 20, "https://xn--bdka7fb.jp/42015.html", "VERIFIED", ""),
    ("タマトア", "タマトア", "P3", 16, "https://xn--bdka7fb.jp/136126.html", "VERIFIED", ""),
    ("シア・カーン", "シアカーン", "P3", 16, "https://xn--bdka7fb.jp/134650.html", "VERIFIED", ""),
    ("ヨロコビ", "ヨロコビ", "P2", 14, "https://xn--bdka7fb.jp/125527.html", "VERIFIED", ""),
    ("カナシミ", "カナシミ", "P2", 14, "https://xn--bdka7fb.jp/125529.html", "VERIFIED", ""),
    ("ハンク", "ハンク", "P2", 14, "https://xn--bdka7fb.jp/125525.html", "VERIFIED", ""),
    ("ラーヤ", "ラーヤ", "P3", 20, "https://xn--bdka7fb.jp/123956.html", "VERIFIED", ""),
    ("シスー", "シスー", "P3", 16, "https://xn--bdka7fb.jp/123958.html", "VERIFIED", ""),
    ("トレメイン夫人", "トレメインフジン", "P2", 14, "https://xn--bdka7fb.jp/122883.html", "VERIFIED", ""),
    ("イズマ", "イズマ", "P2", 14, "https://xn--bdka7fb.jp/122161.html", "VERIFIED", ""),
    ("ルシファー", "ルシファー", "P2", 14, "https://xn--bdka7fb.jp/120722.html", "VERIFIED", ""),
    ("オリバー", "オリバー", "P2", 14, "https://xn--bdka7fb.jp/120724.html", "VERIFIED", ""),
    ("ジョー", "ジョー", "P3", 16, "https://xn--bdka7fb.jp/117259.html", "VERIFIED", ""),
    ("22番", "ニジュウニバン", "P3", 16, "https://xn--bdka7fb.jp/117261.html", "VERIFIED", "ソウルフルワールド"),
    ("エリオット", "エリオット", "P3", 16, "https://xn--bdka7fb.jp/115994.html", "VERIFIED", ""),
    ("フロロー", "フロロー", "P2", 14, "https://xn--bdka7fb.jp/113375.html", "VERIFIED", ""),
    ("ウェンディ", "ウェンディ", "P3", 16, "https://xn--bdka7fb.jp/111838.html", "VERIFIED", ""),
    ("ジョン", "ジョン", "P2", 14, "https://xn--bdka7fb.jp/111841.html", "VERIFIED", ""),
    ("ダンテ", "ダンテ", "P2", 14, "https://xn--bdka7fb.jp/110332.html", "VERIFIED", ""),
    ("ワート", "ワート", "P3", 16, "https://xn--bdka7fb.jp/109311.html", "VERIFIED", ""),
    ("マーリン", "マーリン", "P2", 14, "https://xn--bdka7fb.jp/109313.html", "VERIFIED", ""),
    ("キャンディ大王", "キャンディダイオウ", "P2", 14, "https://xn--bdka7fb.jp/170655.html", "VERIFIED", ""),
    ("チーズ", "チーズ", "P2", 14, "https://xn--bdka7fb.jp/169005.html", "VERIFIED", ""),
    ("ハンス王子", "ハンスオウジ", "P3", 16, "https://xn--bdka7fb.jp/162450.html", "VERIFIED", ""),
    ("マシュマロウ", "マシュマロウ", "P2", 14, "https://xn--bdka7fb.jp/162452.html", "VERIFIED", ""),
    ("ムーシュー", "ムーシュー", "P2", 14, "https://xn--bdka7fb.jp/160879.html", "VERIFIED", ""),
    ("アーシャ", "アーシャ", "P3", 20, "https://xn--bdka7fb.jp/159947.html", "VERIFIED", ""),
    ("エンバー", "エンバー", "P3", 16, "https://xn--bdka7fb.jp/159056.html", "VERIFIED", ""),
    ("ウェイド", "ウェイド", "P3", 16, "https://xn--bdka7fb.jp/159058.html", "VERIFIED", ""),
    ("トランプ兵", "トランプヘイ", "P2", 14, "https://xn--bdka7fb.jp/152054.html", "VERIFIED", ""),
    ("ヴァネッサ", "ヴァネッサ", "P3", 16, "https://xn--bdka7fb.jp/151489.html", "VERIFIED", ""),
    ("カジモド", "カジモド", "P2", 14, "https://xn--bdka7fb.jp/146337.html", "VERIFIED", ""),
    ("正直ジョン", "ショウジキジョン", "P3", 16, "https://xn--bdka7fb.jp/145379.html", "VERIFIED", ""),
    ("ギデオン", "ギデオン", "P2", 14, "https://xn--bdka7fb.jp/145382.html", "VERIFIED", ""),
    ("ミラベル", "ミラベル", "P3", 16, "https://xn--bdka7fb.jp/143266.html", "VERIFIED", ""),
    ("イサベラ", "イサベラ", "P2", 14, "https://xn--bdka7fb.jp/143268.html", "VERIFIED", ""),
    ("ルカ・パグーロ", "ルカパグーロ", "P3", 16, "https://xn--bdka7fb.jp/142185.html", "VERIFIED", ""),
    ("メイリン・リー", "メイリンリー", "P3", 16, "https://xn--bdka7fb.jp/142187.html", "VERIFIED", ""),
    ("ソックス", "ソックス", "P2", 14, "https://xn--bdka7fb.jp/141795.html", "VERIFIED", ""),
    ("モンストロ", "モンストロ", "P3", 20, "https://xn--bdka7fb.jp/140042.html", "VERIFIED", ""),
    ("ロズ", "ロズ", "P3", 16, "https://xn--bdka7fb.jp/137143.html", "VERIFIED", ""),
    ("変身の達人マウイ", "ヘンシンノタツジンマウイ", "P3", 20, "https://xn--bdka7fb.jp/202710.html", "VERIFIED", "変身ツム"),
    ("ジム・ホーキンス", "ジムホーキンス", "P3", 16, "https://xn--bdka7fb.jp/202713.html", "VERIFIED", "2026-06新ツム"),
    ("ヒーロースタイルスティッチ", "ヒーロースタイルスティッチ", "P3", 20, "https://xn--bdka7fb.jp/202429.html", "VERIFIED", "変身ツム"),
    ("ダグ", "ダグ", "P3", 16, "https://xn--bdka7fb.jp/202431.html", "VERIFIED", ""),
    ("ゲイリー", "ゲイリー", "P2", 14, "https://xn--bdka7fb.jp/184389.html", "VERIFIED", ""),
    ("イイナー", "イイナー", "P2", 14, "https://xn--bdka7fb.jp/178197.html", "VERIFIED", ""),
    ("シンパイ", "シンパイ", "P3", 16, "https://xn--bdka7fb.jp/178195.html", "VERIFIED", ""),
    ("コトゥ", "コトゥ", "P3", 16, "https://xn--bdka7fb.jp/177882.html", "VERIFIED", ""),
    ("スミー", "スミー", "P3", 16, "https://xn--bdka7fb.jp/175171.html", "VERIFIED", ""),
    ("イアーゴ", "イアーゴ", "P2", 14, "https://xn--bdka7fb.jp/174305.html", "VERIFIED", ""),
    ("魔法の洞窟", "マホウノドウクツ", "P3", 20, "https://xn--bdka7fb.jp/173877.html", "VERIFIED", ""),
    ("カー", "カー", "P2", 14, "https://xn--bdka7fb.jp/172869.html", "VERIFIED", ""),
    ("ベルウェザー副市長", "ベルウェザーフクシチョウ", "P2", 14, "https://xn--bdka7fb.jp/171043.html", "VERIFIED", ""),
    ("エルネスト・デラクルス", "エルネストデラクルス", "P3", 16, "https://xn--bdka7fb.jp/170651.html", "VERIFIED", ""),
]

# ===================================================================
# ハピネスBOX (常駐14体, SL3 max)
# ===================================================================
# (name, yomi, game8_url, bdka7fb_url, verified, note)
# 全14体: SL1→2=2, SL2→3=4 (合計7でスキルマ)
HAPPINESS_TSUMS = [
    ("ミッキー", "ミッキー", "https://game8.jp/tsumtsum/18404", "https://xn--bdka7fb.jp/2450.html", "VERIFIED", ""),
    ("ミニー", "ミニー", "https://game8.jp/tsumtsum/18292", "https://xn--bdka7fb.jp/4118.html", "VERIFIED", ""),
    ("ドナルド", "ドナルド", "https://game8.jp/tsumtsum/18716", "https://xn--bdka7fb.jp/2637.html", "VERIFIED", ""),
    ("デイジー", "デイジー", "https://game8.jp/tsumtsum/19294", "https://xn--bdka7fb.jp/4160.html", "VERIFIED", ""),
    ("グーフィー", "グーフィー", "https://game8.jp/tsumtsum/19380", "https://xn--bdka7fb.jp/4343.html", "VERIFIED", ""),
    ("プルート", "プルート", "https://game8.jp/tsumtsum/18405", "https://xn--bdka7fb.jp/4311.html", "VERIFIED", ""),
    ("チップ", "チップ", "https://game8.jp/tsumtsum/18263", "https://xn--bdka7fb.jp/3840.html", "VERIFIED", ""),
    ("デール", "デール", "https://game8.jp/tsumtsum/26343", "https://xn--bdka7fb.jp/3856.html", "VERIFIED", ""),
    ("プー", "プー", "https://game8.jp/tsumtsum/18403", "https://xn--bdka7fb.jp/4206.html", "VERIFIED", "くまのプーさん"),
    ("ピグレット", "ピグレット", "https://game8.jp/tsumtsum/18641", "https://xn--bdka7fb.jp/2079.html", "VERIFIED", ""),
    ("ティガー", "ティガー", "https://game8.jp/tsumtsum/19412", "https://xn--bdka7fb.jp/4452.html", "VERIFIED", ""),
    ("イーヨー", "イーヨー", "https://game8.jp/tsumtsum/18237", "https://xn--bdka7fb.jp/3982.html", "VERIFIED", ""),
    ("クリストファーロビン", "クリストファーロビン", "https://game8.jp/tsumtsum/19424", "https://xn--bdka7fb.jp/4838.html", "VERIFIED", ""),
    ("ルー", "ルー", "https://game8.jp/tsumtsum/19411", "https://xn--bdka7fb.jp/4730.html", "VERIFIED", ""),
]

# ===================================================================
# Output path
# ===================================================================
OUT = r"C:\Users\kingu\OneDrive\デスクトップ\ツムツムマスター\data\ガチャデータ.xlsx"
CUSTOM_ORDER_PATH = r"C:\Users\kingu\OneDrive\デスクトップ\ツムツムマスター\data\custom_order.json"


def load_custom_order():
    import json, os
    if not os.path.exists(CUSTOM_ORDER_PATH):
        return None
    with open(CUSTOM_ORDER_PATH, encoding="utf-8") as f:
        return json.load(f)


# ---------- styling helpers ----------
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
PARTIAL_FILL = PatternFill("solid", fgColor="FFF3CD")
LIMITED_FILL = PatternFill("solid", fgColor="FFC7CE")
URL_FONT = Font(color="0563C1", underline="single")


def style_header(ws, headers):
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def add_premium_sheet(wb, title, tsums, custom_order_key_fmt, custom_order, limited_sort_numbers):
    """Add a プレミアム-style sheet (SL1→2 .. SL5→6, 12 cols).

    tsums: list of (name, yomi, pattern, sl5to6, source_url, verified, note)
    custom_order_key_fmt: lambda orig_id -> custom_order key string (e.g. lambda i: f"p-{i:03d}")
    custom_order: dict | None
    limited_sort_numbers: set[int] — sort_no が ここに入ったら赤背景
    """
    ws = wb.create_sheet(title)
    headers = [
        "No", "ID", "ツム名", "成長パターン",
        "SL1→2", "SL2→3", "SL3→4", "SL4→5", "SL5→6",
        "スキルマ合計", "信頼度", "出典URL",
    ]
    style_header(ws, headers)

    with_orig_id = [(i + 1, *t) for i, t in enumerate(tsums)]
    if custom_order:
        sorted_tsums = sorted(
            with_orig_id,
            key=lambda r: custom_order.get(custom_order_key_fmt(r[0]), 9999),
        )
    else:
        sorted_tsums = sorted(with_orig_id, key=lambda r: r[2])  # by yomi

    for sort_no, (orig_id, name, _yomi, pattern, sl56, src, verified, _note) in enumerate(sorted_tsums, start=1):
        base = PATTERN_BASES[pattern]
        c12, c23, c34, c45 = base
        row = [
            sort_no, custom_order_key_fmt(orig_id), name, pattern,
            c12, c23, c34, c45, sl56,
            None, verified, src,
        ]
        ws.append(row)
        r = sort_no + 1
        is_limited = sort_no in limited_sort_numbers

        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=2).alignment = CENTER
        ws.cell(row=r, column=3).alignment = LEFT
        ws.cell(row=r, column=4).alignment = CENTER
        for c in range(5, 10):
            ws.cell(row=r, column=c).alignment = CENTER
        ws.cell(row=r, column=10).value = (
            f"=IF(COUNT(E{r}:I{r})=5,SUM(E{r}:I{r}),\"\")"
        )
        ws.cell(row=r, column=10).alignment = CENTER
        ws.cell(row=r, column=11).alignment = CENTER
        if verified == "PARTIAL":
            ws.cell(row=r, column=11).fill = PARTIAL_FILL
        url_cell = ws.cell(row=r, column=12)
        url_cell.alignment = LEFT
        if src:
            url_cell.hyperlink = src
            url_cell.font = URL_FONT
        for c in range(1, 13):
            ws.cell(row=r, column=c).border = BORDER
            if is_limited:
                ws.cell(row=r, column=c).fill = LIMITED_FILL
        if is_limited and src:
            url_cell.font = Font(color="0563C1", underline="single")

    widths = {
        1: 5, 2: 8, 3: 22, 4: 12,
        5: 7, 6: 7, 7: 7, 8: 7, 9: 7,
        10: 11, 11: 9, 12: 42,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:L{len(sorted_tsums) + 1}"
    return ws


def add_plus_sheet(wb, title, tsums):
    """プレミアムBOX+ シート。Premium と同じ列だが、出典が game8 と xn--bdka7fb の2URL。
    入力: (name, yomi, pattern, sl5to6, game8_url, bdka7fb_url, verified, note)
    """
    ws = wb.create_sheet(title)
    headers = [
        "No", "ID", "ツム名", "成長パターン",
        "SL1→2", "SL2→3", "SL3→4", "SL4→5", "SL5→6",
        "スキルマ合計", "信頼度", "出典URL(game8)", "出典URL(攻略ガイド)", "備考",
    ]
    style_header(ws, headers)

    for sort_no, (name, _yomi, pattern, sl56, game8_url, bdka_url, verified, note) in enumerate(tsums, start=1):
        base = PATTERN_BASES[pattern]
        c12, c23, c34, c45 = base
        row = [
            sort_no, f"pp-{sort_no:03d}", name, pattern,
            c12, c23, c34, c45, sl56,
            None, verified, game8_url, bdka_url, note,
        ]
        ws.append(row)
        r = sort_no + 1
        is_limited = name in LIMITED_NAMES
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=2).alignment = CENTER
        ws.cell(row=r, column=3).alignment = LEFT
        ws.cell(row=r, column=4).alignment = CENTER
        for c in range(5, 10):
            ws.cell(row=r, column=c).alignment = CENTER
        ws.cell(row=r, column=10).value = (
            f"=IF(COUNT(E{r}:I{r})=5,SUM(E{r}:I{r}),\"\")"
        )
        ws.cell(row=r, column=10).alignment = CENTER
        ws.cell(row=r, column=11).alignment = CENTER
        if verified == "PARTIAL":
            ws.cell(row=r, column=11).fill = PARTIAL_FILL
        for url_col, url in ((12, game8_url), (13, bdka_url)):
            cell = ws.cell(row=r, column=url_col)
            cell.alignment = LEFT
            if url:
                cell.hyperlink = url
                cell.font = URL_FONT
        ws.cell(row=r, column=14).alignment = LEFT
        for c in range(1, 15):
            ws.cell(row=r, column=c).border = BORDER
            if is_limited:
                ws.cell(row=r, column=c).fill = LIMITED_FILL

    widths = {
        1: 5, 2: 8, 3: 26, 4: 12,
        5: 7, 6: 7, 7: 7, 8: 7, 9: 7,
        10: 11, 11: 9, 12: 42, 13: 42, 14: 28,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:N{len(tsums) + 1}"
    return ws


def add_happiness_sheet(wb, title, tsums):
    """ハピネスBOXシート。SL1→2 と SL2→3 だけ。
    入力: (name, yomi, game8_url, bdka7fb_url, verified, note)
    """
    ws = wb.create_sheet(title)
    headers = [
        "No", "ID", "ツム名", "成長パターン",
        "SL1→2", "SL2→3",
        "スキルマ合計", "信頼度", "出典URL(game8)", "出典URL(攻略ガイド)", "備考",
    ]
    style_header(ws, headers)

    c12, c23 = HAPPINESS_GROWTH

    for sort_no, (name, _yomi, game8_url, bdka_url, verified, note) in enumerate(tsums, start=1):
        row = [
            sort_no, f"h-{sort_no:03d}", name, HAPPINESS_PATTERN_LABEL,
            c12, c23,
            None, verified, game8_url, bdka_url, note,
        ]
        ws.append(row)
        r = sort_no + 1
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=2).alignment = CENTER
        ws.cell(row=r, column=3).alignment = LEFT
        ws.cell(row=r, column=4).alignment = CENTER
        ws.cell(row=r, column=5).alignment = CENTER
        ws.cell(row=r, column=6).alignment = CENTER
        ws.cell(row=r, column=7).value = (
            f"=IF(COUNT(E{r}:F{r})=2,SUM(E{r}:F{r}),\"\")"
        )
        ws.cell(row=r, column=7).alignment = CENTER
        ws.cell(row=r, column=8).alignment = CENTER
        if verified == "PARTIAL":
            ws.cell(row=r, column=8).fill = PARTIAL_FILL
        for url_col, url in ((9, game8_url), (10, bdka_url)):
            cell = ws.cell(row=r, column=url_col)
            cell.alignment = LEFT
            if url:
                cell.hyperlink = url
                cell.font = URL_FONT
        ws.cell(row=r, column=11).alignment = LEFT
        for c in range(1, 12):
            ws.cell(row=r, column=c).border = BORDER

    widths = {
        1: 5, 2: 8, 3: 22, 4: 12,
        5: 7, 6: 7,
        7: 11, 8: 9, 9: 42, 10: 42, 11: 28,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:K{len(tsums) + 1}"
    return ws


def add_readme_sheet(wb):
    readme = wb.create_sheet("README")
    readme.column_dimensions["A"].width = 95
    readme["A1"] = "ガチャデータ マスター (3 BOX分のスキル成長コスト・出典付き)"
    readme["A1"].font = Font(bold=True, size=14)

    pre_v = sum(1 for t in PREMIUM_TSUMS if t[5] == "VERIFIED")
    pre_p = sum(1 for t in PREMIUM_TSUMS if t[5] == "PARTIAL")
    plus_v = sum(1 for t in PLUS_TSUMS if t[6] == "VERIFIED")
    plus_p = sum(1 for t in PLUS_TSUMS if t[6] == "PARTIAL")
    hap_v = sum(1 for t in HAPPINESS_TSUMS if t[4] == "VERIFIED")
    hap_p = sum(1 for t in HAPPINESS_TSUMS if t[4] == "PARTIAL")

    notes = [
        "",
        "■ シート構成",
        "・プレミアムBOX＋  ... 24体 (2026-06ラインナップ)",
        "・プレミアムBOX    ... 150体 (常駐)",
        "・ハピネスBOX      ... 14体 (常駐, SL3 max)",
        "",
        "■ データ品質",
        f"・プレミアムBOX＋: VERIFIED {plus_v}体 / PARTIAL {plus_p}体 / 合計 {len(PLUS_TSUMS)}体",
        f"・プレミアムBOX  : VERIFIED {pre_v}体 / PARTIAL {pre_p}体 / 合計 {len(PREMIUM_TSUMS)}体",
        f"・ハピネスBOX    : VERIFIED {hap_v}体 / PARTIAL {hap_p}体 / 合計 {len(HAPPINESS_TSUMS)}体",
        "・全件、game8.jp と xn--bdka7fb.jp の個別ツム評価ページから一次確認",
        "・出典URLは各行に明記。クリックで該当ページに遷移",
        "",
        "■ 並び順 (プレミアムBOX)",
        "・No 列: ユーザー指定のカスタム順 (data/custom_order.json)",
        "・ID 列: ゲーム内インフォメーション順の固定ID (p-001..p-150)",
        "",
        "■ 行の色",
        "・赤背景: 期間限定キャラ (プレミアムBOXシートの No.147/149/150)",
        "・黄背景 (信頼度列のみ): PARTIAL — 出典に異本あり、または片サイトのみ確認",
        "",
        "■ このシートの値",
        "・SL1→2 ... SL5→6: そのレベルに上げるためにガチャで重複させる必要があるツム数",
        "・スキルマ合計: 全段階の合計 = スキルマまでに必要な被りツム総数",
        "・「スキル発動に必要なツム数」(SL1〜6) とは別概念",
        "",
        "■ 成長パターン",
        "・P1: 2-3-4-6-X  (2015年3月以前の常駐ツム)",
        "・P2: 1-2-4-7-X  (中堅〜21型ツム)",
        "・P3: 1-2-4-8-X  (晩成型・近年の重め成長)",
        "・PS: 1-2-2-3-X  (特殊。ハム専用、合計11と最軽量)",
        "・HP: 2-4        (ハピネスBOX専用。SL3 maxなので2段階のみ。合計6被り=7体でMAX)",
        "・X = SL5→6 (ツム個別、3〜21の範囲)",
        "・プレミアムBOX＋ の「+ツム」は全員 P3 (1,2,4,8) + sl5to6=20 で統一",
        "",
        "■ 出典の主要サイト",
        "・xn--bdka7fb.jp (LINEディズニー ツムツム攻略・裏ワザ徹底ガイド) ← 個別評価ページ",
        "・game8.jp/tsumtsum/* (ゲームエイト個別評価ページ)",
        "・https://game8.jp/tsumtsum/23195 (スキル6必要数集約表)",
        "",
        "■ 編集方針",
        "・本ファイルは app/src/data/premium.ts, happiness.ts と内容を同期させる (順序は違ってOK)",
        "・新ツム追加・値更新時は ID と ソースURLを必ず記録",
    ]
    for line in notes:
        readme.append([line])


def main():
    custom_order = load_custom_order()

    wb = Workbook()
    # Remove default sheet — we'll add explicit ones in the desired order
    default = wb.active
    wb.remove(default)

    # 1. プレミアムBOX＋
    add_plus_sheet(wb, "プレミアムBOX＋", PLUS_TSUMS)

    # 2. プレミアムBOX
    add_premium_sheet(
        wb,
        "プレミアムBOX",
        PREMIUM_TSUMS,
        lambda i: f"p-{i:03d}",
        custom_order,
        limited_sort_numbers={147, 149, 150},
    )

    # 3. ハピネスBOX
    add_happiness_sheet(wb, "ハピネスBOX", HAPPINESS_TSUMS)

    # 4. README
    add_readme_sheet(wb)

    wb.save(OUT)
    print(f"Saved: {OUT}")
    print(f"  プレミアムBOX+ : {len(PLUS_TSUMS)} tsums")
    print(f"  プレミアムBOX  : {len(PREMIUM_TSUMS)} tsums")
    print(f"  ハピネスBOX    : {len(HAPPINESS_TSUMS)} tsums")


if __name__ == "__main__":
    main()
